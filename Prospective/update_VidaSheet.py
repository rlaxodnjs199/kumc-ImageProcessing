# Usage
# - python update_datasheet.py
# Dependency
# - VidaDatasheet.xlsx
# Global Variable -----------------------------------------------------------
XLSX_PATH = r"E:\common\Taewon\oneDrive\OneDrive - University of Kansas Medical Center\VidaSheet.xlsx"
# ---------------------------------------------------------------------------
import pandas as pd
from pandas import DataFrame
from openpyxl import load_workbook


def _find_first_case_id_to_update(vida_datasheet_df: DataFrame) -> int:
    return int(vida_datasheet_df.iloc[-1]["VidaCaseID"]) + 1


def _constuct_df_to_append(
    vida_start_case_id: int, vida_dashboard_df: DataFrame
) -> DataFrame:
    vida_dashboard_df_filtered = vida_dashboard_df[
        vida_dashboard_df["Case ID"] >= vida_start_case_id
    ]
    rows_to_append = []
    for _, df_row in vida_dashboard_df_filtered.iterrows():
        try:
            row = {}
            row["Proj"] = ""
            row["Subj"] = df_row["Patient Name"]
            row["VidaCaseID"] = df_row["Case ID"]
            row["VidaBy"] = ""
            row["MRN"] = ""
            row["Vida Progress"] = df_row["Process status"]
            row["Progress"] = ""
            row["ScanDate"] = int(df_row["Acquisition Date"].strftime("%Y%m%d"))
            row["IN/EX"] = df_row["Scan Type"]
            row["CT Protocol"] = df_row["Series Desc"]
            row["Disease"] = ""
            row["SliceThickness_mm"] = df_row["Slice Thickness"]
            row["ScannerVender"] = df_row["Scanner"]
            row["ScannerModel"] = df_row["Scanner Model"]
            row["Kernel"] = df_row["Kernel"]
            row["Comments"] = ""
            row["VIDA path"] = ""
            row["Report path"] = ""

            rows_to_append.append(row)

        except:
            print("constructing dataframe failed at VidaCaseID: ", df_row["Case ID"])
            continue

    return pd.DataFrame(rows_to_append)


def _append_df_to_excel(
    filename, df, sheet_name="Sheet1", startrow=None, **to_excel_kwargs
):
    with pd.ExcelWriter(filename, engine="openpyxl", mode="a") as writer:
        writer.book = load_workbook(filename)
        startrow = writer.book[sheet_name].max_row
        if startrow is None:
            startrow = 0
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)


def update_vida_datasheet(vida_datasheet_df: DataFrame, vida_dashboard_df: DataFrame):
    vida_start_case_id = _find_first_case_id_to_update(vida_datasheet_df)
    print(f"Starting from Case ID: {vida_start_case_id}...")
    df_to_append = _constuct_df_to_append(vida_start_case_id, vida_dashboard_df)
    _append_df_to_excel(XLSX_PATH, df_to_append, header=None, index=False)
    print(f"Update VidaSheet.xlsx: Done")


if __name__ == "__main__":
    vida_datasheet_df = pd.read_excel(XLSX_PATH, sheet_name=0)
    vida_dashboard_df = pd.read_excel(XLSX_PATH, sheet_name=1)
    update_vida_datasheet(vida_datasheet_df, vida_dashboard_df)

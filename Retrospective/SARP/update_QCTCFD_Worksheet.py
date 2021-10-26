from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

VIDA_RESULT_PATH = r"E:\VIDA\VIDAvision2.2"
VIDASHEET_PATH = r"E:\common\Taewon\oneDrive\OneDrive - University of Kansas Medical Center\VidaSheet.xlsx"
MASTERSHEET_PATH = r"P:\IRB_STUDY00144315_AATRU\Jiwoong\SARP\Data\SARP_DCC\SARP_ClusterCT\SARP_I_II_and_III_master_data_report _04112019_DL.xlsx"
QCTCFD_WORKSHEET_PATH = r"E:\common\Taewon\oneDrive\OneDrive - University of Kansas Medical Center\QCTCFD_Worksheet.xlsx"

def extract_subj(subj):
    return subj.split('_')[0]

def extract_FU(subj):
    return int(subj[-1]) - 1

def trim_subj(subj):
    return '-'.join(subj.split('-')[:-1])

def construct_df_QCTCFD_from_VIDA_dashboard(df_QCTCFD: pd.DataFrame, df_VIDA: pd.DataFrame):
    for _, row in df_VIDA.iterrows():
        if row['CT Protocol'].split(' ')[0] == 'INSPIRATION':
            df_QCTCFD.loc[(df_QCTCFD['Subj'] == row['Subj']) & (df_QCTCFD['Date'] == row['ScanDate']), 'VIDA_IN'] = 'O'
            df_QCTCFD.loc[(df_QCTCFD['Subj'] == row['Subj']) & (df_QCTCFD['Date'] == row['ScanDate']), 'VIDA_IN_Path'] = f'{VIDA_RESULT_PATH}\{row.VidaCaseID}'
            df_QCTCFD.loc[(df_QCTCFD['Subj'] == row['Subj']) & (df_QCTCFD['Date'] == row['ScanDate']), 'VIDA_IN_At'] = 'KUMC'
        elif row['CT Protocol'].split(' ')[0] == 'EXPIRATION':
            df_QCTCFD.loc[(df_QCTCFD['Subj'] == row['Subj']) & (df_QCTCFD['Date'] == row['ScanDate']), 'VIDA_EX'] = 'O'
            df_QCTCFD.loc[(df_QCTCFD['Subj'] == row['Subj']) & (df_QCTCFD['Date'] == row['ScanDate']), 'VIDA_EX_Path'] = f'{VIDA_RESULT_PATH}\{row.VidaCaseID}'
            df_QCTCFD.loc[(df_QCTCFD['Subj'] == row['Subj']) & (df_QCTCFD['Date'] == row['ScanDate']), 'VIDA_EX_At'] = 'KUMC'

def append_df_to_excel(
    filename,
    df,
    sheet_name="Sheet1",
    startrow=None,
    **to_excel_kwargs
):
    with pd.ExcelWriter(filename, engine="openpyxl", mode="a", date_format='m/d/yyyy', datetime_format='m/d/yyyy') as writer:
        writer.book = load_workbook(filename)
        startrow = writer.book[sheet_name].max_row
        if startrow is None:
            startrow = 0
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

# DF: SARP3 available in VIDA_machine:/e/jchoi4/ImageData/SARP/VIDA_20201102-03_SARP_VIDA
def update_QCTCFD_WORKSHEET_columns(filename,
    df,
    sheet_name="Sheet1",
    **to_excel_kwargs
):
    with pd.ExcelWriter(filename, engine="openpyxl", mode="a", date_format='m/d/yyyy', datetime_format='m/d/yyyy') as writer:
        writer.book = load_workbook(filename)
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        df.to_excel(writer, sheet_name, startrow=0, **to_excel_kwargs)

# DF: SARP3 imported to VIDA
df_VIDA = pd.read_excel(VIDASHEET_PATH)
df_SARP3 = df_VIDA.loc[df_VIDA['Subj'].str.contains('80-'), :]
df_SARP3['Subj'] = df_SARP3['Subj'].apply(extract_subj)
df_SARP3['ScanDate'] = pd.to_datetime(df_SARP3['ScanDate'], format='%Y%m%d')
# df_SARP3_unique_group = df_SARP3.groupby(['Subj', 'ScanDate']).size().reset_index().rename(columns={0:'count'})

# DF: SARP3 available in B2
# df_1 = pd.read_csv(r'C:\Users\tkim3\Documents\Codes\ImageProcessing\Scripts\Temp\Subj_Date_SARP3_EX0.csv', names=['Subj', 'ScanDate'])
# df_2 = pd.read_csv(r'C:\Users\tkim3\Documents\Codes\ImageProcessing\Scripts\Temp\Subj_Date_SARP3_EX1.csv', names=['Subj', 'ScanDate'])
# df_3 = pd.read_csv(r'C:\Users\tkim3\Documents\Codes\ImageProcessing\Scripts\Temp\Subj_Date_SARP3_IN0.csv', names=['Subj', 'ScanDate'])
# df_4 = pd.read_csv(r'C:\Users\tkim3\Documents\Codes\ImageProcessing\Scripts\Temp\Subj_Date_SARP3_IN1.csv', names=['Subj', 'ScanDate'])
# df_SARP3_B2 = pd.concat([df_1, df_2, df_3, df_4]).reset_index(drop=True)
# print(df_SARP3_B2)
# df_SARP3_B2_unique = df_SARP3_B2.groupby(['Subj', 'ScanDate']).size().reset_index().rename(columns={0:'count'})
# # print(df_SARP3_B2_unique)

# QCTCFD_Worksheet DF
df_MASTER = pd.read_excel(MASTERSHEET_PATH)
df_QCTCFD = pd.read_excel(QCTCFD_WORKSHEET_PATH)
df_MASTER_SARP3 = df_MASTER.loc[df_MASTER['EDF_ID'].str.contains('80-'), :][['EDF_ID', 'study_date']]
df_MASTER_SARP3['FU'] = df_MASTER_SARP3['EDF_ID'].apply(extract_FU)
df_MASTER_SARP3['EDF_ID'] = df_MASTER_SARP3['EDF_ID'].apply(trim_subj)
df_MASTER_SARP3['study_date'] = pd.to_datetime(df_MASTER_SARP3['study_date'], format='%Y%m%d')
df_MASTER_SARP3.rename(columns={'EDF_ID':'Subj', 'study_date': 'Date'}, inplace=True)
df_MASTER_SARP3_unique = df_MASTER_SARP3.groupby(['Subj', 'Date', 'FU']).size().reset_index().rename(columns={0:'count'}).drop(columns='count')
df_MASTER_SARP3_unique['Proj'] = 'SARP3'
df_QCTCFD = pd.concat([df_QCTCFD, df_MASTER_SARP3_unique]).fillna('')

construct_df_QCTCFD_from_VIDA_dashboard(df_QCTCFD, df_SARP3)
# append_df_to_excel(QCTCFD_WORKSHEET_PATH, df_QCTCFD, header=None, index=False)
update_QCTCFD_WORKSHEET_columns(QCTCFD_WORKSHEET_PATH, df_QCTCFD, index=False)

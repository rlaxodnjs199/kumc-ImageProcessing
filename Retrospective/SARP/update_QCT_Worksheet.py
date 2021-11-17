from typing import Dict
import pandas as pd
import numpy as np
from openpyxl import load_workbook

VIDA_RESULT_PATH = r"E:\VIDA\VIDAvision2.2"
VIDASHEET_PATH = r"E:\common\Taewon\oneDrive\OneDrive - University of Kansas Medical Center\VidaSheet.xlsx"
MASTERSHEET_PATH = r"P:\IRB_STUDY00144315_AATRU\Jiwoong\SARP\Data\SARP_DCC\SARP_ClusterCT\SARP_I_II_and_III_master_data_report _04112019_DL.xlsx"
QCTCFD_WORKSHEET_PATH = r"E:\common\Taewon\oneDrive\OneDrive - University of Kansas Medical Center\QCTCFD_Worksheet.xlsx"


def append_query_condition_for_the_DF_sent_to_B2() -> str:
    return "VIDA_IN == 'O' and VIDA_EX == 'O' and (FU == 1 or FU == 0)"


def extract_subj(subj) -> str:
    return subj.split("_")[0]


def extract_FU(subj) -> int:
    return int(subj[-1]) - 1


def trim_subj(subj) -> str:
    return "-".join(subj.split("-")[:-1])


def construct_df_QCTCFD_from_VIDA_dashboard(
    df_QCTCFD: pd.DataFrame, df_VIDA: pd.DataFrame
):
    for _, row in df_VIDA.iterrows():
        if row["CT Protocol"].split(" ")[0] == "INSPIRATION":
            df_QCTCFD_filtered = df_QCTCFD.query(
                "Subj == @row.Subj and Date == @row.ScanDate"
            )
            df_QCTCFD.loc[
                df_QCTCFD_filtered.index,
                "DCM_IN",
            ] = "O"
            df_QCTCFD.loc[
                df_QCTCFD_filtered.index,
                "VIDA_IN",
            ] = "O"
            df_QCTCFD.loc[
                df_QCTCFD_filtered.index,
                "VIDA_IN_Path",
            ] = f"{VIDA_RESULT_PATH}\{row.VidaCaseID}"
            df_QCTCFD.loc[
                df_QCTCFD_filtered.index,
                "VIDA_IN_At",
            ] = "KUMC"
        elif row["CT Protocol"].split(" ")[0] == "EXPIRATION":
            df_QCTCFD_filtered = df_QCTCFD.query(
                "Subj == @row.Subj and Date == @row.ScanDate"
            )
            df_QCTCFD.loc[
                df_QCTCFD_filtered.index,
                "DCM_EX",
            ] = "O"
            df_QCTCFD.loc[
                df_QCTCFD_filtered.index,
                "VIDA_EX",
            ] = "O"
            df_QCTCFD.loc[
                df_QCTCFD_filtered.index,
                "VIDA_EX_Path",
            ] = f"{VIDA_RESULT_PATH}\{row.VidaCaseID}"
            df_QCTCFD.loc[
                df_QCTCFD_filtered.index,
                "VIDA_EX_At",
            ] = "KUMC"

    return


def update_df_QCTCFD_sent_to_remote_host(df_QCTCFD: pd.DataFrame):
    condition: str = append_query_condition_for_the_DF_sent_to_B2()
    df_QCTCFD.loc[df_QCTCFD.query(condition).index, "IR"] = "O"

    return


def append_df_to_excel(
    filename, df, sheet_name="Sheet1", startrow=None, **to_excel_kwargs
):
    with pd.ExcelWriter(
        filename,
        engine="openpyxl",
        mode="a",
        date_format="m/d/yyyy",
        datetime_format="yyyymmdd",
    ) as writer:
        writer.book = load_workbook(filename)
        startrow = writer.book[sheet_name].max_row
        if startrow is None:
            startrow = 0
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    return


def update_QCTCFD_WORKSHEET(filename, df, sheet_name="Sheet1", **to_excel_kwargs):
    with pd.ExcelWriter(
        filename,
        engine="xlsxwriter",
        date_format="m/d/yyyy",
        datetime_format="yyyymmdd",
    ) as writer:
        df.to_excel(writer, sheet_name=sheet_name, **to_excel_kwargs)
        writer.save()

    return


def initialize_df_QCTCFD(df_MASTER_SARP3_unique: pd.DataFrame) -> pd.DataFrame:
    df_QCTCFD = pd.DataFrame(
        columns=[
            "Proj",
            "Subj",
            "Date",
            "FU",
            "DCM_IN",
            "DCM_EX",
            "VIDA_IN",
            "VIDA_EX",
            "VIDA_IN_Path",
            "VIDA_EX_Path",
            "VIDA_IN_At",
            "VIDA_EX_At",
            "IR",
            "QCT",
            "PostIR",
            "CFD1D",
            "CFD3D",
            "Ptcl1D",
            "Ptcl3D",
        ]
    )
    df_QCTCFD = pd.concat([df_QCTCFD, df_MASTER_SARP3_unique]).fillna("")

    return df_QCTCFD


def prepare_SARP3_list_from_master_datasheet(
    master_data_sheet_path=MASTERSHEET_PATH,
) -> pd.DataFrame:
    df_MASTER = pd.read_excel(master_data_sheet_path)
    df_MASTER_SARP3 = df_MASTER.loc[df_MASTER["EDF_ID"].str.contains("80-"), :][
        ["EDF_ID", "study_date"]
    ]
    df_MASTER_SARP3["FU"] = df_MASTER_SARP3["EDF_ID"].apply(extract_FU)
    df_MASTER_SARP3["EDF_ID"] = df_MASTER_SARP3["EDF_ID"].apply(trim_subj)
    df_MASTER_SARP3["study_date"] = pd.to_datetime(
        df_MASTER_SARP3["study_date"], format="%Y%m%d"
    )
    df_MASTER_SARP3.rename(
        columns={"EDF_ID": "Subj", "study_date": "Date"}, inplace=True
    )
    df_MASTER_SARP3_unique = (
        df_MASTER_SARP3.groupby(["Subj", "Date", "FU"])
        .size()
        .reset_index()
        .rename(columns={0: "count"})
        .drop(columns="count")
    )
    df_MASTER_SARP3_unique["Proj"] = "SARP"

    return df_MASTER_SARP3_unique


def prepare_SARP_list_from_master_datasheet(df_MASTER) -> pd.DataFrame:
    df_MASTER_SARP = df_MASTER.loc[df_MASTER["EDF_ID"].str.len() == 5, :][
        ["EDF_ID", "project", "scan_type", "study_date"]
    ]

    return df_MASTER_SARP


def find_SARP3_missing_data(df_MASTER, df_QCT_worksheet) -> pd.DataFrame:
    df_MASTER_SARP3 = df_MASTER.loc[df_MASTER["EDF_ID"].str.contains("80-"), :]

    df_SARP3_missing = pd.DataFrame(
        columns=["EDF_ID", "project", "scan_type", "study_date"]
    )

    for _, row in df_QCT_worksheet.iterrows():
        if row["DCM_IN"] != "O":
            df_SARP3_missing = df_SARP3_missing.append(
                search_SARP3_entry_from_master_datasheet(
                    df_MASTER_SARP3, row["Subj"], row["Date"], row["FU"], "IN"
                ),
                ignore_index=True,
            )
        if row["DCM_EX"] != "O":
            df_SARP3_missing = df_SARP3_missing.append(
                search_SARP3_entry_from_master_datasheet(
                    df_MASTER_SARP3, row["Subj"], row["Date"], row["FU"], "EX"
                ),
                ignore_index=True,
            )

    return df_SARP3_missing


def search_SARP3_entry_from_master_datasheet(
    df_MASTER_SARP3, subj, date, fu, in_or_ex
) -> Dict:
    fu = str(fu + 1)
    SARP3_edf_id = f"{subj}-V{fu}"
    if in_or_ex == "IN":
        df_query_result = df_MASTER_SARP3.query(
            "EDF_ID == @SARP3_edf_id and scan_type == 'Inspiratory'"
        )
    else:
        df_query_result = df_MASTER_SARP3.query(
            "EDF_ID == @SARP3_edf_id and scan_type == 'Expiratory'"
        )
    row = df_query_result.loc[
        df_query_result.index, ["EDF_ID", "project", "scan_type", "study_date"]
    ]

    return row


def prepare_SARP3_df_imported_to_VIDA(VIDA_sheet_path=VIDASHEET_PATH) -> pd.DataFrame:
    df_VIDA = pd.read_excel(VIDA_sheet_path)
    df_SARP3 = df_VIDA.loc[df_VIDA["Subj"].str.contains("80-"), :]
    df_SARP3["Subj"] = df_SARP3["Subj"].apply(extract_subj)
    df_SARP3["ScanDate"] = pd.to_datetime(df_SARP3["ScanDate"], format="%Y%m%d")
    # df_SARP3_unique_group = df_SARP3.groupby(['Subj', 'ScanDate']).size().reset_index().rename(columns={0:'count'})

    return df_SARP3


def prepare_SARP3_df_available_in_B2():
    # df_1 = pd.read_csv(r'C:\Users\tkim3\Documents\Codes\ImageProcessing\Scripts\Temp\Subj_Date_SARP3_EX0.csv', names=['Subj', 'ScanDate'])
    # df_2 = pd.read_csv(r'C:\Users\tkim3\Documents\Codes\ImageProcessing\Scripts\Temp\Subj_Date_SARP3_EX1.csv', names=['Subj', 'ScanDate'])
    # df_3 = pd.read_csv(r'C:\Users\tkim3\Documents\Codes\ImageProcessing\Scripts\Temp\Subj_Date_SARP3_IN0.csv', names=['Subj', 'ScanDate'])
    # df_4 = pd.read_csv(r'C:\Users\tkim3\Documents\Codes\ImageProcessing\Scripts\Temp\Subj_Date_SARP3_IN1.csv', names=['Subj', 'ScanDate'])
    # df_SARP3_B2 = pd.concat([df_1, df_2, df_3, df_4]).reset_index(drop=True)
    # print(df_SARP3_B2)
    # df_SARP3_B2_unique = df_SARP3_B2.groupby(['Subj', 'ScanDate']).size().reset_index().rename(columns={0:'count'})
    # # print(df_SARP3_B2_unique)

    return


if __name__ == "__main__":
    df_MASTER = pd.read_excel(MASTERSHEET_PATH)
    df_QCT_worksheet = pd.read_excel(QCTCFD_WORKSHEET_PATH)

    # df_SARP3 = prepare_SARP3_df_imported_to_VIDA()
    # df_QCTCFD = initialize_df_QCTCFD(prepare_SARP3_list_from_master_datasheet())
    # construct_df_QCTCFD_from_VIDA_dashboard(df_QCTCFD, df_SARP3)
    # update_df_QCTCFD_sent_to_remote_host(df_QCTCFD)
    # update_QCTCFD_WORKSHEET(QCTCFD_WORKSHEET_PATH, df_QCTCFD, index=False)

    # Find SARP 1,2,3 missing data: 2021/11/15
    df_SARP12_missing_data = prepare_SARP_list_from_master_datasheet(df_MASTER)
    df_SARP_missing_data = df_SARP12_missing_data.append(
        find_SARP3_missing_data(df_MASTER, df_QCT_worksheet)
    )
    df_SARP_missing_data.to_excel(
        r"C:\Users\tkim3\Documents\Codes\ImageProcessing\Output\SARP_missing_data.xlsx",
        index=False,
    )

# Usage
# - python update_datasheet.py
# Dependency
# - Datasheet.xlsx
# - COVID_CTs_20210513CarissaWalter_0608jc.xlsx

import pandas as pd
import os
from openpyxl import load_workbook

PROJ = "C19"
DISEASE = "COVID-19"

EXCEL_PATH = "Data/Datasheet/DataSheet.xlsx"
MRN_CTDATE_PATH = "Data/CarissaWalter/COVID_CTs_20210513CarissaWalter_0608jc.xlsx"
VIDA_PATH_PREFIX = ""
DCM_PATH_PREFIX = ""

datasheet_df = pd.read_excel(EXCEL_PATH, sheet_name=0)
vida_dashboard_df = pd.read_excel(EXCEL_PATH, sheet_name=1)
carissa_df = pd.read_excel(MRN_CTDATE_PATH, header=9)

start_vida_case_number = int(datasheet_df.iloc[-1]["VidaCaseID"])
vida_dashboard_df = vida_dashboard_df[
    vida_dashboard_df["Case ID"] > start_vida_case_number
]
merged_df = pd.merge(
    vida_dashboard_df, carissa_df, how="inner", left_on="Patient ID", right_on="Subj"
)


def construct_new_rows(merged_df):
    rows_to_append = []
    for _, df_row in merged_df.iterrows():
        row = {}
        accession_number = df_row["Accession Number"]
        subj_id = accession_number.split("_")[0]
        img_id = accession_number.split("_")[1]

        row["Proj"] = PROJ
        row["Subj"] = subj_id
        row["VidaCaseID"] = df_row["Case ID"]
        row["VidaBy"] = ""
        row["MRN"] = df_row["mrn"]
        row["Vida Progress"] = df_row["Process status"]
        row["Progress"] = "DL Segmentation Done"
        row["ScanDate"] = (
            int(df_row["date"].strftime("%Y%m%d"))
            if type(df_row["date"]) == pd.Timestamp
            else ""
        )
        row["IN/EX"] = img_id
        row["CT Protocol"] = df_row["Series Desc"]
        row["Disease"] = DISEASE
        row["SliceThickness_mm"] = df_row["Slice Thickness"]
        row["ScannerVender"] = df_row["Scanner"]
        row["ScannerModel"] = df_row["Scanner Model"]
        row["Kernel"] = df_row["Kernel"]
        row["Comments"] = ""
        row["VIDA path"] = VIDA_PATH_PREFIX
        row["Report path"] = DCM_PATH_PREFIX

        rows_to_append.append(row)

    return pd.DataFrame(rows_to_append)


def append_df_to_excel(
    filename,
    df,
    sheet_name="Sheet1",
    startrow=None,
    truncate_sheet=False,
    **to_excel_kwargs
):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs
        )
        return

    # ignore [engine] parameter if it was passed
    if "engine" in to_excel_kwargs:
        to_excel_kwargs.pop("engine")

    writer = pd.ExcelWriter(filename, engine="openpyxl", mode="a")

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


new_rows_df = construct_new_rows(merged_df)
append_df_to_excel(
    EXCEL_PATH, new_rows_df, sheet_name="Sheet1", header=None, index=False
)

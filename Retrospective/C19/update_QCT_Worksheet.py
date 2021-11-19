from typing import Dict
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import load_workbook

CARISSA_WALTER_DATASHEET_PATH = r"C:\Users\tkim3\Documents\Codes\ImageProcessing\Data\CarissaWalter\COVID_CTs_20210513CarissaWalter_0608jc.xlsx"
VIDASHEET_PATH = r"E:\common\Taewon\oneDrive\OneDrive - University of Kansas Medical Center\VidaSheet.xlsx"
QCT_WORKSHEET_PATH = r"E:\common\Taewon\oneDrive\OneDrive - University of Kansas Medical Center\QCT_Worksheet.xlsx"
VIDA_RESULT_PATH = r"E:\VIDA\VIDAvision2.2"

Carissa_df = pd.read_excel(
    CARISSA_WALTER_DATASHEET_PATH, header=9, usecols="A,C,I,J"
).sort_values(by=["mrn", "Time"])
VIDAsheet_df = pd.read_excel(VIDASHEET_PATH)
QCTworksheet_df = pd.read_excel(QCT_WORKSHEET_PATH)


def initialize_C19_base_df() -> pd.DataFrame:
    C19_df = pd.DataFrame(
        columns=[
            "Proj",
            "Subj",
            "Date",
            "FU",
            "DCM_IN",
            "DCM_EX",
            "VIDA_IN",
            "VIDA_EX",
            "VIDA_IN_Path",
            "VIDA_EX_Path",
            "VIDA_IN_At",
            "VIDA_EX_At",
            "IR",
            "QCT",
            "PostIR",
            "CFD1D",
            "CFD3D",
            "Ptcl1D",
            "Ptcl3D",
        ]
    )

    return C19_df


def date_to_str(date: datetime) -> str:
    return date.strftime("%Y%m%d")


def float_to_str(float: float) -> str:
    return str(float)


def initialize_C19_df_from_Carissa_df(
    C19_base_df: pd.DataFrame, Carissa_df: pd.DataFrame
) -> pd.DataFrame:
    C19_from_Carissa_df = Carissa_df.drop(columns="mrn")
    C19_from_Carissa_df["Proj"] = "C19"
    C19_from_Carissa_df = C19_from_Carissa_df.rename(
        columns={"Time": "FU", "date": "Date"}
    )
    C19_from_Carissa_df.dropna(subset=["Date"], inplace=True)
    C19_from_Carissa_df["Date"] = C19_from_Carissa_df["Date"].apply(date_to_str)
    C19_df = pd.concat([C19_base_df, C19_from_Carissa_df]).fillna("")

    return C19_df


def preprocess_VIDAsheet_df(VIDAsheet_df: pd.DataFrame) -> pd.DataFrame:
    VIDAsheet_df.dropna(subset=["ScanDate"], inplace=True)
    VIDAsheet_df["ScanDate"] = VIDAsheet_df["ScanDate"].astype(int).astype(str)

    return VIDAsheet_df


def update_C19_df_from_VIDAsheet_df(
    C19_total_df: pd.DataFrame, VIDAsheet_df: pd.DataFrame
) -> pd.DataFrame:
    for index, row in C19_total_df.iterrows():
        case = VIDAsheet_df.query("Subj == @row.Subj and ScanDate == @row.Date")
        if not case.empty:
            C19_total_df.loc[index, "DCM_IN"] = "O"
            if case.iloc[0].Progress == "Done":
                C19_total_df.loc[index, "VIDA_IN"] = "O"
                C19_total_df.loc[
                    index, "VIDA_IN_Path"
                ] = f"{VIDA_RESULT_PATH}\{case.iloc[0].VidaCaseID}"
                C19_total_df.loc[index, "VIDA_IN_At"] = "KUMC"

    return C19_total_df


def append_df_to_excel(
    filename, df, sheet_name="C19", startrow=None, **to_excel_kwargs
):
    with pd.ExcelWriter(
        filename,
        engine="openpyxl",
        mode="a",
        date_format="m/d/yyyy",
        datetime_format="yyyymmdd",
    ) as writer:
        writer.book = load_workbook(filename)
        startrow = writer.book[sheet_name].max_row
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    return


if __name__ == "__main__":
    C19_base_df = initialize_C19_base_df()
    C19_total_df = initialize_C19_df_from_Carissa_df(C19_base_df, Carissa_df)
    VIDAsheet_df = preprocess_VIDAsheet_df(VIDAsheet_df)
    C19_total_df = update_C19_df_from_VIDAsheet_df(C19_total_df, VIDAsheet_df)

    append_df_to_excel(QCT_WORKSHEET_PATH, C19_total_df, index=False)
